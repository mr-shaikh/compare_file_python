import openpyxl.styles as styles
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from tkinter import filedialog, messagebox
import codecs
import os
from datetime import date
from openpyxl.styles import Font, PatternFill
import openpyxl


def download_file(folder1_path, folder2_path, file_list):
    if file_list is None or file_list.size() == 0:
        print("No files to download.")
        return

    # Select a destination folder to save the downloaded files
    destination_folder = filedialog.askdirectory()

    if not destination_folder:
        print("No destination folder selected.")
        return

    # Create a new workbook to store the error files
    error_book = openpyxl.Workbook()
    red_font = Font(color="00FF0000")
    blue_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')

    # Add a summary sheet for file names and statuses
    summary_sheet = error_book.active
    summary_sheet.title = "Summary"

    # Create a style for the header cells
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')

    # Append the header row with colorful and bold cells
    header_row = ["File Name", "Status"]
    summary_sheet.append(header_row)
    for cell in summary_sheet[1]:
        cell.font = header_style
        cell.fill = header_fill

    for i in range(file_list.size()):
        filename = file_list.get(i)
        file1_path = os.path.join(folder1_path, filename)
        file2_path = os.path.join(folder2_path, filename)

        if not os.path.isfile(file1_path):
            # If file1 is missing, add it to the summary sheet as missing
            summary_sheet.append([filename, "Missing"])
            continue  # Skip further processing for this file

        try:
            with codecs.open(file1_path, "rb", encoding="shift-jis") as file1, \
                    codecs.open(file2_path, "rb", encoding="shift-jis") as file2:
                file1_contents = file1.read()
                file2_contents = file2.read()
        except UnicodeDecodeError:
            print(f"Error reading file {filename}: UnicodeDecodeError")
            continue

        if "Missing" in filename or "Extra" in filename:
            # If the file name contains "Missing" or "Extra", skip it
            continue
        else:
            # If the file has differences, add it to the error workbook

            # Create a new sheet for the error file
            sheet = error_book.create_sheet(title=filename)

            # Write the error file header row with light blue background color
            header = ["Row Number", "BELC", "VEBUIN"]
            sheet.append(header)
            for cell in sheet[1]:
                cell.fill = blue_fill

            # Set font color to red for the error file name
            sheet["A1"].font = red_font

            # Parse the error rows from file1 and file2 contents and write to the sheet
            file1_lines = file1_contents.split("\n")
            file2_lines = file2_contents.split("\n")
            for i, (line1, line2) in enumerate(zip(file1_lines, file2_lines), start=1):
                if line1 != line2:
                    row = [i, line1, line2]
                    sheet.append(row)

                    # Check if the characters are different from BELC's row
                    belc_row = file1_lines[i-1]  # Assuming BELC's row is in file1_lines
                    for j, (char1, char2, belc_char) in enumerate(zip(line1, line2, belc_row), start=1):
                        if char1 != belc_char:
                            cell = sheet.cell(row=i+1, column=j+1)
                            cell.font = Font(color="00FF0000")  # Set font color to red
                            if j == 3:  # Assuming "VEBUIN" column is in column 3
                                vebuin_cell = sheet.cell(row=i+1, column=j+1)
                                vebuin_cell.font = Font(color="00FF0000")  # Set font color to red
            # Add file name and status to the summary sheet
            summary_sheet.append([filename, "Differences"])

            # Change the color of different characters in the "VEBUIN" column
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if cell.value != belc_row[cell.row - 2]:
                        cell.font = Font(color="00FF0000")  # Set font color to red

    # Save the error workbook
    today = date.today().strftime("%Y-%m-%d")
    error_file_name = f"Error_Files_{today}.xlsx"
    error_file_path = os.path.join(destination_folder, error_file_name)
    error_book.save(error_file_path)
    print(f"Error files saved as {error_file_path}")